#!/usr/bin/env python3
"""
상세기능설계서 엑셀 생성 스크립트
WBS JSON 데이터를 기반으로 프로세스별 상세기능설계서를 생성합니다.

사용법:
  python create_dev_spec.py wbs.json --output 상세기능설계서.xlsx --project-code PC_PLC
"""

import argparse
import json
import sys
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl --break-system-packages")
    sys.exit(1)

# 스타일 정의
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LABEL_FONT = Font(name="맑은 고딕", size=10, bold=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def generate_process_id(project_code: str, group_idx: int, process_idx: int) -> str:
    """프로세스 ID를 생성합니다. 예: PC_PLC_0101"""
    return f"{project_code}_{group_idx:02d}{process_idx:02d}"


def group_wbs_into_processes(wbs_items: list[dict]) -> list[dict]:
    """WBS 항목을 프로세스(Level 3 단위)로 그룹화합니다."""
    processes = []
    current_l2 = ""
    group_idx = 0
    process_idx = 0

    for item in wbs_items:
        l2 = item.get("level2", "")
        l3 = item.get("level3", "")
        l4 = item.get("level4", "")

        if l2 != current_l2:
            current_l2 = l2
            group_idx += 1
            process_idx = 0

        # Level 3가 새 프로세스의 시작
        existing = next((p for p in processes if p["level3"] == l3 and p["level2"] == l2), None)

        if existing:
            # 기존 프로세스에 Step 추가
            existing["steps"].append({
                "level4": l4,
                "description": item.get("description", ""),
                "repos": item.get("repos", ""),
                "notes": item.get("notes", ""),
            })
        else:
            process_idx += 1
            processes.append({
                "level1": item.get("level1", ""),
                "level2": l2,
                "level3": l3,
                "group_idx": group_idx,
                "process_idx": process_idx,
                "steps": [{
                    "level4": l4,
                    "description": item.get("description", ""),
                    "repos": item.get("repos", ""),
                    "notes": item.get("notes", ""),
                }],
            })

    return processes


def estimate_effort(description: str) -> float:
    """설명 기반으로 공수를 추정합니다."""
    desc_lower = description.lower()
    # 키워드 기반 추정
    if any(kw in desc_lower for kw in ["ai", "모델", "학습", "스트리밍", "실시간"]):
        return 3.0
    elif any(kw in desc_lower for kw in ["api 연동", "oauth", "외부", "배치"]):
        return 2.0
    elif any(kw in desc_lower for kw in ["crud", "조회", "목록", "설정", "토글"]):
        return 0.5
    elif any(kw in desc_lower for kw in ["등록", "수정", "삭제", "관리"]):
        return 1.0
    return 1.0


def infer_role(repos: str) -> str:
    """레포명에서 역할을 추론합니다."""
    if not repos:
        return "FS"
    repos_lower = repos.lower()
    if "frontend" in repos_lower and "backend" in repos_lower:
        return "FS"
    elif "frontend" in repos_lower:
        return "FE"
    elif "llm" in repos_lower or "ai" in repos_lower:
        return "AI"
    else:
        return "BE"


def create_toc_sheet(wb: Workbook, processes: list[dict], project_code: str):
    """목차 시트를 생성합니다."""
    ws = wb.active
    ws.title = "목차"

    headers = ["No", "프로세스 ID", "Level 1\n(업무영역)", "Level 2\n(업무기능)",
               "프로세스명", "담당자", "Step 수", "예상 공수(일)"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 35

    for row_idx, proc in enumerate(processes, 2):
        proc_id = generate_process_id(project_code, proc["group_idx"], proc["process_idx"])
        total_effort = sum(estimate_effort(s.get("description", "")) for s in proc["steps"])

        row_data = [
            row_idx - 1,
            proc_id,
            proc["level1"],
            proc["level2"],
            proc["level3"],
            "",  # 담당자 (비워둠)
            len(proc["steps"]),
            total_effort,
        ]

        fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") \
            if (row_idx - 2) % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    col_widths = [5, 16, 15, 15, 25, 12, 8, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:H{len(processes) + 1}"


def create_process_sheet(wb: Workbook, proc: dict, project_code: str):
    """프로세스별 시트를 생성합니다."""
    proc_id = generate_process_id(project_code, proc["group_idx"], proc["process_idx"])
    sheet_name = proc_id[-4:]  # 시트명은 4자리 번호 (길이 제한)
    ws = wb.create_sheet(sheet_name)

    # 상단: 프로세스 정보
    info_labels = ["프로세스 ID", "프로세스명", "Input", "Output", "설명", "요구사항 참조", "담당자"]
    info_values = [
        proc_id,
        proc["level3"],
        "",  # Input (사용자가 채움)
        "",  # Output (사용자가 채움)
        f"{proc['level1']} > {proc['level2']} > {proc['level3']}",
        "",  # 요구사항 참조 (사용자가 채움)
        "",  # 담당자 (사용자가 채움)
    ]

    for row_idx, (label, value) in enumerate(zip(info_labels, info_values), 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.fill = LABEL_FILL
        label_cell.font = LABEL_FONT
        label_cell.border = THIN_BORDER

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = DATA_FONT
        value_cell.border = THIN_BORDER
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)

    # 빈 행
    start_row = len(info_labels) + 2

    # 하단: Step 상세 테이블
    step_headers = ["Step 번호", "역할", "Step명", "수행방안", "레포",
                    "개발 설계 상세", "완료 조건", "예상 공수(일)"]

    for col_idx, header in enumerate(step_headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[start_row].height = 30

    for step_idx, step in enumerate(proc["steps"], 1):
        row = start_row + step_idx
        effort = estimate_effort(step.get("description", ""))
        role = infer_role(step.get("repos", ""))
        step_num = f"{proc_id}_{step_idx:02d}"

        step_data = [
            step_num,
            role,
            step.get("level4", ""),
            step.get("description", ""),
            step.get("repos", ""),
            "",  # 개발 설계 상세 (사용자가 채움)
            "",  # 완료 조건 (사용자가 채움)
            effort,
        ]

        fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") \
            if step_idx % 2 == 1 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col_idx, val in enumerate(step_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    col_widths = [16, 6, 20, 35, 20, 35, 25, 10]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    parser = argparse.ArgumentParser(description="상세기능설계서 엑셀 생성")
    parser.add_argument("input_file", help="WBS JSON 파일")
    parser.add_argument("--output", "-o", default="상세기능설계서.xlsx", help="출력 엑셀 파일")
    parser.add_argument("--project-code", "-p", default="PC_PLC", help="프로젝트 코드 (기본: PC_PLC)")
    args = parser.parse_args()

    with open(args.input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wbs_items = data.get("wbs", data.get("items", []))
    processes = group_wbs_into_processes(wbs_items)

    wb = Workbook()
    create_toc_sheet(wb, processes, args.project_code)

    for proc in processes:
        create_process_sheet(wb, proc, args.project_code)

    # 작성 가이드 시트
    guide_ws = wb.create_sheet("작성 가이드")
    guide_data = [
        ["필드", "설명", "작성 예시"],
        ["Step 번호", "프로세스ID_순번 형식", "PC_PLC_0101_01"],
        ["역할", "BE/FE/AI/FS 중 하나", "BE"],
        ["Step명", "구현할 단위 기능명", "이메일 로그인 API"],
        ["수행방안", "구현 방법 설명", "JWT 기반 인증, bcrypt 비밀번호 해싱"],
        ["레포", "구현 대상 레포", "backend-auth"],
        ["개발 설계 상세", "상세 구현 내용", "POST /auth/login endpoint 구현"],
        ["완료 조건", "완료로 인정되는 조건", "로그인 API 정상 응답, 테스트 통과"],
        ["예상 공수", "일 단위 (S:0.5, M:1, L:2, XL:3+)", "1.0"],
    ]
    for row_idx, row in enumerate(guide_data, 1):
        for col_idx, val in enumerate(row, 1):
            cell = guide_ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = DATA_FONT
            cell.border = THIN_BORDER
    for col_idx, w in enumerate([18, 30, 40], 1):
        guide_ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(args.output)
    print(f"상세기능설계서 생성 완료: {args.output} ({len(processes)}개 프로세스)")


if __name__ == "__main__":
    main()
