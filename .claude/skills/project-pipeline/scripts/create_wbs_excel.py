#!/usr/bin/env python3
"""
업무기능분해도(WBS) 엑셀 생성 스크립트
JSON 형식의 WBS 데이터를 엑셀 파일로 변환합니다.

사용법:
  python create_wbs_excel.py wbs.json --output 업무기능분해도.xlsx
"""

import argparse
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl --break-system-packages")
    sys.exit(1)


# 스타일 정의
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
ODD_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
EVEN_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def create_wbs_excel(wbs_data: list[dict], output_path: str):
    """WBS 데이터로 엑셀 파일을 생성합니다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "업무기능분해도"

    # 헤더
    headers = ["No", "Level 1\n(업무영역)", "Level 2\n(업무기능)", "Level 3\n(세부기능)",
               "Level 4\n(단위기능)", "기능 설명", "관련 레포", "비고"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 헤더 행 높이
    ws.row_dimensions[1].height = 35

    # 데이터 입력
    prev_l1 = prev_l2 = prev_l3 = ""

    for row_idx, item in enumerate(wbs_data, 2):
        no = item.get("no", row_idx - 1)
        l1 = item.get("level1", "")
        l2 = item.get("level2", "")
        l3 = item.get("level3", "")
        l4 = item.get("level4", "")
        desc = item.get("description", "")
        repos = item.get("repos", "")
        notes = item.get("notes", "")

        # 셀 병합을 위해 이전 값과 같으면 빈 문자열
        display_l1 = l1 if l1 != prev_l1 else ""
        display_l2 = l2 if l2 != prev_l2 or l1 != prev_l1 else ""
        display_l3 = l3 if l3 != prev_l3 or l2 != prev_l2 else ""

        row_data = [no, display_l1, display_l2, display_l3, l4, desc, repos, notes]
        fill = ODD_FILL if (row_idx - 2) % 2 == 0 else EVEN_FILL

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

        prev_l1, prev_l2, prev_l3 = l1, l2, l3

    # 열 너비 설정
    col_widths = [5, 15, 15, 15, 20, 40, 25, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 필터 추가
    ws.auto_filter.ref = f"A1:H{len(wbs_data) + 1}"

    # 작성 가이드 시트
    guide_ws = wb.create_sheet("작성 가이드")
    guide_data = [
        ["레벨", "설명", "기준"],
        ["Level 1 (업무영역)", "최상위 업무 영역", "요구사항정의서의 대분류에 대응"],
        ["Level 2 (업무기능)", "주요 기능 그룹", "요구사항정의서의 중분류에 대응"],
        ["Level 3 (세부기능)", "구현 모듈 수준", "설계 시 모듈/서비스 단위"],
        ["Level 4 (단위기능)", "최소 개발 단위", "하나의 PR로 완료 가능한 수준"],
    ]
    for row_idx, row in enumerate(guide_data, 1):
        for col_idx, val in enumerate(row, 1):
            cell = guide_ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            else:
                cell.font = DATA_FONT
            cell.border = THIN_BORDER

    for col_idx, width in enumerate([25, 25, 45], 1):
        guide_ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)
    print(f"업무기능분해도 생성 완료: {output_path} ({len(wbs_data)}개 항목)")


def main():
    parser = argparse.ArgumentParser(description="업무기능분해도 엑셀 생성")
    parser.add_argument("input_file", help="WBS JSON 파일")
    parser.add_argument("--output", "-o", default="업무기능분해도.xlsx", help="출력 엑셀 파일")
    args = parser.parse_args()

    with open(args.input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wbs_items = data.get("wbs", data.get("items", []))
    create_wbs_excel(wbs_items, args.output)


if __name__ == "__main__":
    main()
