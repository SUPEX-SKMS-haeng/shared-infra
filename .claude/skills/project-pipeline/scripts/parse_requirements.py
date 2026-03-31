#!/usr/bin/env python3
"""
요구사항정의서 파싱 스크립트
엑셀, 마크다운, CSV 형식의 요구사항정의서를 JSON으로 변환합니다.

사용법:
  python parse_requirements.py <input_file> --output requirements.json
"""

import argparse
import json
import re
import sys
import os

# 컬럼명 매핑 (한글 → 영문 키)
COLUMN_MAPPINGS = {
    "req_id": ["요구사항 ID", "요구사항ID", "ID", "번호", "No"],
    "req_type": ["요구사항구분", "구분", "유형", "Type"],
    "category_l1": ["대분류", "업무영역", "L1", "대분류(업무영역)"],
    "category_l2": ["중분류", "업무기능", "L2", "중분류(업무기능)"],
    "category_l3": ["소분류", "세부기능", "L3", "소분류(세부기능)"],
    "req_name": ["요구사항명", "기능명", "제목", "Name"],
    "description": ["요구사항 설명", "설명", "상세설명", "Description"],
    "priority": ["우선순위", "Priority"],
    "status": ["상태", "Status"],
    "mgmt_type": ["관리구분", "관리", "Management"],
    "updated_at": ["최종변경일", "변경일", "수정일"],
    "effort_mm": ["개발M/M", "공수", "M/M", "인월", "예상공수", "개발 M/M"],
    "effort_by_role": ["역할별 공수", "역할별공수", "Role Effort"],
    "notes": ["비고", "참고사항", "Notes", "Remarks"],
}


def find_column_key(header: str) -> str | None:
    """헤더 텍스트에서 영문 키를 찾습니다."""
    header = header.strip()
    for key, aliases in COLUMN_MAPPINGS.items():
        for alias in aliases:
            if alias.lower() == header.lower():
                return key
    return None


def parse_excel(file_path: str) -> list[dict]:
    """엑셀 파일을 파싱합니다."""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl이 필요합니다: pip install openpyxl --break-system-packages")
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active  # 첫 번째 시트 사용

    # 헤더 행 찾기 (첫 10행 내에서 '요구사항' 또는 'ID' 포함 행)
    header_row = None
    headers = []
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
        for val in row_values:
            if find_column_key(val):
                header_row = row_idx
                headers = row_values
                break
        if header_row:
            break

    if not header_row:
        print(f"헤더 행을 찾을 수 없습니다. 첫 10행 내에 요구사항 관련 컬럼명이 있어야 합니다.")
        sys.exit(1)

    # 컬럼 매핑
    col_map = {}
    for col_idx, header in enumerate(headers):
        key = find_column_key(header)
        if key:
            col_map[key] = col_idx

    print(f"  헤더 행: {header_row}")
    print(f"  매핑된 컬럼: {list(col_map.keys())}")

    # 데이터 파싱
    requirements = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]

        # 빈 행 스킵
        if not any(row_values):
            continue

        req = {}
        for key, col_idx in col_map.items():
            if col_idx < len(row_values):
                req[key] = row_values[col_idx]

        # req_id 또는 req_name이 있어야 유효한 행
        if req.get("req_id") or req.get("req_name"):
            requirements.append(req)

    return requirements


def parse_markdown(file_path: str) -> list[dict]:
    """마크다운 파일을 파싱합니다."""
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()

    requirements = []
    # FR-XXX 또는 NFR-XXX 패턴 찾기
    pattern = r"###\s+((?:FR|NFR)-\d+):\s*(.+?)(?=\n###|\n##|\Z)"
    matches = re.finditer(pattern, content, re.DOTALL)

    for match in matches:
        req_id = match.group(1)
        block = match.group(2).strip()

        req = {"req_id": req_id}
        req["req_type"] = "기능" if req_id.startswith("FR") else "비기능"

        # 제목 추출 (첫 줄)
        lines = block.split("\n")
        req["req_name"] = lines[0].strip() if lines else ""

        # 필드 추출
        for line in lines:
            line = line.strip()
            if line.startswith("- **설명**:"):
                req["description"] = line.replace("- **설명**:", "").strip()
            elif line.startswith("- **우선순위**:"):
                req["priority"] = line.replace("- **우선순위**:", "").strip()
            elif line.startswith("- **관련 레포**:"):
                req["notes"] = line.replace("- **관련 레포**:", "").strip()
            elif line.startswith("- **비고**:"):
                notes = line.replace("- **비고**:", "").strip()
                req["notes"] = (req.get("notes", "") + " " + notes).strip()

        requirements.append(req)

    return requirements


def parse_csv(file_path: str) -> list[dict]:
    """CSV 파일을 파싱합니다."""
    import csv

    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader)

        col_map = {}
        for col_idx, header in enumerate(headers):
            key = find_column_key(header.strip())
            if key:
                col_map[key] = col_idx

        requirements = []
        for row in reader:
            if not any(row):
                continue
            req = {}
            for key, col_idx in col_map.items():
                if col_idx < len(row):
                    req[key] = row[col_idx].strip()
            if req.get("req_id") or req.get("req_name"):
                requirements.append(req)

    return requirements


def main():
    parser = argparse.ArgumentParser(description="요구사항정의서 파싱")
    parser.add_argument("input_file", help="입력 파일 경로 (.xlsx, .md, .csv)")
    parser.add_argument("--output", "-o", default="requirements.json", help="출력 JSON 파일")
    args = parser.parse_args()

    ext = os.path.splitext(args.input_file)[1].lower()
    print(f"파싱 중: {args.input_file} ({ext})")

    if ext in (".xlsx", ".xls"):
        requirements = parse_excel(args.input_file)
    elif ext == ".md":
        requirements = parse_markdown(args.input_file)
    elif ext == ".csv":
        requirements = parse_csv(args.input_file)
    else:
        print(f"지원하지 않는 형식: {ext}")
        print("지원 형식: .xlsx, .xls, .md, .csv")
        sys.exit(1)

    print(f"파싱 완료: {len(requirements)}개 요구사항")

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump({"requirements": requirements, "total": len(requirements)}, f, ensure_ascii=False, indent=2)

    print(f"저장됨: {args.output}")


if __name__ == "__main__":
    main()
