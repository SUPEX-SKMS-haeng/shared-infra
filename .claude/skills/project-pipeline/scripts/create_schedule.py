#!/usr/bin/env python3
"""
일정표 엑셀 생성 스크립트
태스크 목록과 팀원 정보를 기반으로 스프린트 일정표를 자동 생성합니다.

사용법:
  python create_schedule.py tasks.json --team-info team.json --output 일정표.xlsx
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl --break-system-packages")
    sys.exit(1)

HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
TASK_FILLS = {
    "BE": PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid"),
    "FE": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "AI": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "FS": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
}
WEEKEND_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def is_workday(date: datetime, holidays: list[str] = None) -> bool:
    """주말과 공휴일을 제외한 근무일인지 확인합니다."""
    if date.weekday() >= 5:  # 토, 일
        return False
    if holidays:
        date_str = date.strftime("%Y-%m-%d")
        for h in holidays:
            if date_str == h or date_str.replace("-", "/").endswith(h.replace("-", "/")):
                return False
    return True


def get_workdays(start: datetime, count: float, holidays: list[str] = None, allocation: float = 1.0) -> tuple:
    """시작일로부터 count 근무일 후의 종료일을 구합니다."""
    actual_days = count / allocation if allocation > 0 else count
    actual_days = max(0.5, actual_days)

    current = start
    days_remaining = actual_days

    while days_remaining > 0:
        if is_workday(current, holidays):
            days_remaining -= 1
        if days_remaining > 0:
            current += timedelta(days=1)

    return current


def topological_sort(tasks: list[dict]) -> list[dict]:
    """의존성 기반 위상 정렬합니다."""
    task_map = {t["id"]: t for t in tasks}
    visited = set()
    result = []

    def visit(task_id):
        if task_id in visited:
            return
        visited.add(task_id)
        task = task_map.get(task_id)
        if task:
            for dep in task.get("depends_on", []):
                if dep in task_map:
                    visit(dep)
            result.append(task)

    for task in tasks:
        visit(task["id"])

    return result


def assign_schedule(tasks: list[dict], team: list[dict], config: dict) -> list[dict]:
    """태스크에 일정을 배정합니다."""
    start_date = datetime.strptime(config.get("start_date", "2026-04-01"), "%Y-%m-%d")
    sprint_days = config.get("sprint_days", 10)
    buffer_rate = config.get("buffer_rate", 0.2)
    holidays = config.get("holidays", [])

    # 팀원별 정보 매핑
    team_map = {}
    for member in team:
        team_map[member["name"]] = {
            "role": member.get("role", "FS"),
            "allocation": member.get("allocation", 1.0),
            "next_available": start_date,
            "absences": member.get("absences", []),
        }

    sorted_tasks = topological_sort(tasks)
    scheduled = []
    task_end_map = {}  # task_id → end_date

    for task in sorted_tasks:
        effort = task.get("effort", 1.0)
        effort_with_buffer = effort * (1 + buffer_rate)

        # 시작일 결정: 의존성 + 담당자 가용일
        earliest_start = start_date
        for dep_id in task.get("depends_on", []):
            if dep_id in task_end_map:
                dep_end = task_end_map[dep_id]
                next_day = dep_end + timedelta(days=1)
                while not is_workday(next_day, holidays):
                    next_day += timedelta(days=1)
                if next_day > earliest_start:
                    earliest_start = next_day

        assignee = task.get("assignee", "")
        allocation = 1.0
        if assignee and assignee in team_map:
            member = team_map[assignee]
            allocation = member["allocation"]
            if member["next_available"] > earliest_start:
                earliest_start = member["next_available"]

        # 시작일이 근무일인지 확인
        while not is_workday(earliest_start, holidays):
            earliest_start += timedelta(days=1)

        # 종료일 계산
        end_date = get_workdays(earliest_start, effort_with_buffer, holidays, allocation)

        # 담당자 가용일 업데이트
        if assignee and assignee in team_map:
            next_avail = end_date + timedelta(days=1)
            while not is_workday(next_avail, holidays):
                next_avail += timedelta(days=1)
            team_map[assignee]["next_available"] = next_avail

        task_end_map[task["id"]] = end_date

        scheduled.append({
            **task,
            "start_date": earliest_start.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "effort_with_buffer": round(effort_with_buffer, 1),
        })

    return scheduled


def create_schedule_excel(scheduled_tasks: list[dict], team: list[dict],
                          config: dict, output_path: str):
    """일정표 엑셀을 생성합니다."""
    wb = Workbook()

    # === 스프린트 계획 시트 ===
    ws = wb.active
    ws.title = "스프린트 계획"

    start_date = datetime.strptime(config.get("start_date", "2026-04-01"), "%Y-%m-%d")
    sprint_days = config.get("sprint_days", 10)

    headers = ["Sprint", "담당자", "역할", "레포", "태스크", "시작일", "종료일", "공수(일)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 스프린트 계산
    for row_idx, task in enumerate(scheduled_tasks, 2):
        task_start = datetime.strptime(task["start_date"], "%Y-%m-%d")
        days_from_start = (task_start - start_date).days
        sprint_num = days_from_start // (sprint_days + 4) + 1  # 주말 포함 근사

        row_data = [
            f"Sprint {sprint_num}",
            task.get("assignee", "미배정"),
            task.get("role", ""),
            task.get("repo", ""),
            task.get("title", task.get("id", "")),
            task["start_date"],
            task["end_date"],
            task.get("effort_with_buffer", task.get("effort", 1.0)),
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP

    col_widths = [10, 10, 6, 18, 30, 12, 12, 8]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # === 팀원별 일정 시트 ===
    gantt_ws = wb.create_sheet("팀원별 일정")

    # 전체 기간 계산
    if scheduled_tasks:
        all_starts = [datetime.strptime(t["start_date"], "%Y-%m-%d") for t in scheduled_tasks]
        all_ends = [datetime.strptime(t["end_date"], "%Y-%m-%d") for t in scheduled_tasks]
        min_date = min(all_starts)
        max_date = max(all_ends)
    else:
        min_date = start_date
        max_date = start_date + timedelta(days=30)

    total_days = (max_date - min_date).days + 1

    # 헤더: 담당자 | 태스크 | 날짜1 | 날짜2 | ...
    gantt_ws.cell(row=1, column=1, value="담당자").fill = HEADER_FILL
    gantt_ws.cell(row=1, column=1).font = HEADER_FONT
    gantt_ws.cell(row=1, column=1).border = THIN_BORDER
    gantt_ws.cell(row=1, column=2, value="태스크").fill = HEADER_FILL
    gantt_ws.cell(row=1, column=2).font = HEADER_FONT
    gantt_ws.cell(row=1, column=2).border = THIN_BORDER

    for day_idx in range(total_days):
        date = min_date + timedelta(days=day_idx)
        col = day_idx + 3
        cell = gantt_ws.cell(row=1, column=col, value=date.strftime("%m/%d"))
        cell.fill = HEADER_FILL if is_workday(date) else WEEKEND_FILL
        cell.font = Font(name="맑은 고딕", size=8, bold=True,
                         color="FFFFFF" if is_workday(date) else "999999")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        gantt_ws.column_dimensions[get_column_letter(col)].width = 4

    gantt_ws.column_dimensions["A"].width = 10
    gantt_ws.column_dimensions["B"].width = 25

    # 태스크별 간트 바
    for row_idx, task in enumerate(scheduled_tasks, 2):
        gantt_ws.cell(row=row_idx, column=1, value=task.get("assignee", "")).font = DATA_FONT
        gantt_ws.cell(row=row_idx, column=1).border = THIN_BORDER
        gantt_ws.cell(row=row_idx, column=2, value=task.get("title", task.get("id", ""))).font = DATA_FONT
        gantt_ws.cell(row=row_idx, column=2).border = THIN_BORDER

        t_start = datetime.strptime(task["start_date"], "%Y-%m-%d")
        t_end = datetime.strptime(task["end_date"], "%Y-%m-%d")
        role = task.get("role", "FS")
        fill = TASK_FILLS.get(role, TASK_FILLS["FS"])

        for day_idx in range(total_days):
            date = min_date + timedelta(days=day_idx)
            col = day_idx + 3
            cell = gantt_ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER

            if t_start <= date <= t_end and is_workday(date):
                cell.fill = fill
                cell.value = "■"
                cell.font = Font(size=8)
                cell.alignment = Alignment(horizontal="center")
            elif not is_workday(date):
                cell.fill = WEEKEND_FILL

    # === 마일스톤 시트 ===
    ms_ws = wb.create_sheet("마일스톤")
    ms_headers = ["마일스톤", "목표 완료일", "관련 태스크", "상태"]
    for col_idx, h in enumerate(ms_headers, 1):
        cell = ms_ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for col_idx, w in enumerate([25, 12, 40, 10], 1):
        ms_ws.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(output_path)
    print(f"일정표 생성 완료: {output_path} ({len(scheduled_tasks)}개 태스크)")


def main():
    parser = argparse.ArgumentParser(description="일정표 엑셀 생성")
    parser.add_argument("input_file", help="태스크 JSON 파일")
    parser.add_argument("--team-info", "-t", help="팀원 정보 JSON 파일")
    parser.add_argument("--output", "-o", default="일정표.xlsx", help="출력 엑셀 파일")
    args = parser.parse_args()

    with open(args.input_file, "r", encoding="utf-8") as f:
        task_data = json.load(f)

    tasks = task_data.get("tasks", [])

    team = []
    config = task_data.get("config", {})
    if args.team_info:
        with open(args.team_info, "r", encoding="utf-8") as f:
            team_data = json.load(f)
        team = team_data.get("team", [])
        config.update(team_data.get("config", {}))

    if not config.get("start_date"):
        config["start_date"] = datetime.now().strftime("%Y-%m-%d")

    scheduled = assign_schedule(tasks, team, config)

    # 리스크 분석
    risks = []
    member_load = defaultdict(float)
    for t in scheduled:
        assignee = t.get("assignee", "미배정")
        member_load[assignee] += t.get("effort_with_buffer", t.get("effort", 1.0))

    sprint_days = config.get("sprint_days", 10)
    for name, load in member_load.items():
        if load > sprint_days * 0.8:
            risks.append(f"{name} 공수 과다: {load:.1f}일/{sprint_days}일")

    if risks:
        print("\n⚠️ 리스크:")
        for r in risks:
            print(f"  - {r}")

    create_schedule_excel(scheduled, team, config, args.output)


if __name__ == "__main__":
    main()
